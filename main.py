import yfinance as yf
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
import os
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                             QHBoxLayout, QLabel, QLineEdit, QPushButton, 
                             QTextEdit, QFileDialog, QMessageBox, QTableWidget,
                             QTableWidgetItem, QAbstractItemView)
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from PyQt5.QtGui import QFont, QPixmap
import sys

def format_currency(value):
    """Format a numeric value as currency (millions)"""
    if value is None or pd.isna(value):
        return 'N/A'
    if abs(value) >= 1000:
        return f"${value/1000000:,.2f}"
    else:
        return f"${value/1000000:,.2f}"

def format_number(value):
    """Format a numeric value (for shares, etc.)"""
    if value is None or pd.isna(value):
        return 'N/A'
    if abs(value) >= 1000:
        return f"{value/1000000:,.2f}"
    else:
        return f"{value/1000000:,.2f}"

def extract_value(df, row_names, column_idx=0):
    """Extract a value from a DataFrame by trying multiple row names"""
    if df.empty:
        return None
    for name in row_names:
        if name in df.index:
            value = df.loc[name].iloc[column_idx]
            if not pd.isna(value):
                return value
    return None

def calculate_net_debt(balance_sheet, column_idx=0):
    """Calculate Net Debt = Total Debt - Cash and Cash Equivalents"""
    if balance_sheet.empty:
        return None
    
    # Try to get Total Debt directly first
    total_debt = None
    if 'Total Debt' in balance_sheet.index:
        total_debt = balance_sheet.loc['Total Debt'].iloc[column_idx]
        if pd.isna(total_debt):
            total_debt = None
    
    # If Total Debt not available, calculate from Long Term + Short Term Debt
    if total_debt is None:
        long_term_debt = None
        short_term_debt = None
        
        # Try different names for long-term debt
        long_term_names = ['Long Term Debt', 'Long Term Debt And Capital Lease Obligation', 'Long Term Debt Noncurrent']
        for name in long_term_names:
            if name in balance_sheet.index:
                long_term_debt = balance_sheet.loc[name].iloc[column_idx]
                if not pd.isna(long_term_debt):
                    break
        
        # Try different names for short-term debt
        short_term_names = ['Current Debt', 'Short Term Debt', 'Short Term Debt And Capital Lease Obligation']
        for name in short_term_names:
            if name in balance_sheet.index:
                short_term_debt = balance_sheet.loc[name].iloc[column_idx]
                if not pd.isna(short_term_debt):
                    break
        
        # Sum them if both available, or use long-term if only that's available
        if long_term_debt is not None:
            total_debt = long_term_debt
            if short_term_debt is not None:
                total_debt += short_term_debt
    
    # Try different names for cash
    cash_names = ['Cash And Cash Equivalents', 'Cash Cash Equivalents And Short Term Investments', 'Cash And Short Term Investments']
    cash = None
    for name in cash_names:
        if name in balance_sheet.index:
            cash = balance_sheet.loc[name].iloc[column_idx]
            if not pd.isna(cash):
                break
    
    if total_debt is None or cash is None:
        return None
    
    return total_debt - cash

def extract_year_from_date(date_str):
    """Extract year from date string"""
    try:
        # Try to parse as YYYY-MM-DD format
        if isinstance(date_str, pd.Timestamp):
            return date_str.year
        elif isinstance(date_str, str):
            # Try to extract year from string (e.g., "2024-12-31" -> 2024)
            if '-' in date_str:
                return int(date_str.split('-')[0])
            # Or if it's just a year
            if len(date_str) == 4:
                return int(date_str)
    except:
        pass
    return None

def get_historical_stock_prices(ticker, start_year=2020):
    """Get historical stock prices on the last trading day of each year"""
    try:
        stock = yf.Ticker(ticker)
        
        # Try to get historical data from start_year to now
        # If start_year is before 2020, try to get from earliest available or 2020
        actual_start_year = max(start_year, 2020) if start_year < 2020 else start_year
        
        hist = stock.history(start=f"{actual_start_year}-01-01", end=None)
        
        if hist.empty:
            # Try to get from beginning if available
            try:
                hist = stock.history(period="max")
            except:
                return {}
        
        if hist.empty:
            return {}
        
        # Group by year and get the last trading day of each year
        prices_by_year = {}
        hist.index = pd.to_datetime(hist.index)
        
        # Get unique years, but filter to only include years from start_year onwards
        years = [y for y in hist.index.year.unique() if y >= start_year]
        
        for year in years:
            year_data = hist[hist.index.year == year]
            if not year_data.empty:
                # Get the last trading day of the year
                last_day = year_data.index.max()
                price = year_data.loc[last_day, 'Close']
                prices_by_year[year] = float(price)
        
        return prices_by_year
    except Exception as e:
        print(f"Error fetching historical stock prices: {e}")
        return {}

def calculate_ev_ebitda_multiple_for_year(stock_price, shares_outstanding, net_debt, ebitda):
    """Calculate EV/EBITDA multiple for a specific year = (Market Cap + Net Debt) / EBITDA
    where Market Cap = Stock Price × Shares Outstanding"""
    try:
        if stock_price is None or shares_outstanding is None or net_debt is None or ebitda is None:
            return None
        if pd.isna(stock_price) or pd.isna(shares_outstanding) or pd.isna(net_debt) or pd.isna(ebitda):
            return None
        if ebitda == 0:
            return None
        
        # Calculate Market Cap = Stock Price × Shares Outstanding
        market_cap = stock_price * shares_outstanding
        
        # Calculate Enterprise Value = Market Cap + Net Debt
        enterprise_value = market_cap + net_debt
        
        # Calculate EV/EBITDA
        multiple = enterprise_value / ebitda
        return multiple
            
    except Exception as e:
        print(f"Error calculating EV/EBITDA multiple: {e}")
        return None

def write_to_excel(years_data, file_path, sheet_name):
    """Write financial data to existing Excel workbook by finding metrics and updating them"""
    try:
        # Load existing workbook
        wb = load_workbook(file_path)
        
        # Get the specified sheet
        if sheet_name not in wb.sheetnames:
            print(f"Error: Sheet '{sheet_name}' not found in workbook.")
            print(f"Available sheets: {', '.join(wb.sheetnames)}")
            return False
        
        ws = wb[sheet_name]
        
        # Define the order of metrics to search for
        metric_order = [
            'Revenue',
            'COGS',
            'Gross Profit',
            'SG&A',
            'D&A',
            'R&D',
            'EBT',
            'EBITDA',
            'Net Income',
            'Free Cash Flow',
            'FCFF',
            'FCFE',
            'LCapex',
            'Net Debt',
            'accounts receivables',
            'prepaid expenses',
            'inventories',
            'accounts payable',
            'TOCA Other Short-Term Liabilities',
            'Taxes',
            'Interest Expense',
            '# FDSO (millions)',
            'multiple',
            'share price'
        ]
        
        # Extract years from data and sort them
        years = []
        for year_data in years_data:
            year = extract_year_from_date(year_data['Year'])
            if year:
                years.append(year)
        
        # Sort years (oldest to newest)
        years = sorted(set(years))
        
        if not years:
            print("Error: No years found in the data.")
            return False
        
        # Find the earliest year from Yahoo Finance data
        yahoo_earliest_year = min(years)
        print(f"Earliest year from Yahoo Finance: {yahoo_earliest_year}")
        
        # Find the earliest year in the Excel file
        max_col = ws.max_column
        max_row = ws.max_row
        excel_years = []
        
        # Scan the entire Excel sheet to find all years (2000-2099)
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value is not None:
                    cell_str = str(cell_value).strip()
                    # Try to extract year from cell (could be "2020", "2020A", etc.)
                    for year in range(2000, 2100):
                        year_str = str(year)
                        if year_str in cell_str:
                            excel_years.append(year)
                            break
        
        if not excel_years:
            print("Error: Could not find any year headers in the Excel sheet.")
            return False
        
        excel_earliest_year = min(excel_years)
        excel_earliest_year_str = str(excel_earliest_year)
        print(f"Earliest year in Excel: {excel_earliest_year}")
        
        # Find the column for the earliest year in Excel
        excel_earliest_year_col = None
        excel_earliest_year_row = None
        
        # Search for the earliest year column in Excel
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value is not None:
                    cell_str = str(cell_value).strip()
                    if excel_earliest_year_str in cell_str:
                        excel_earliest_year_col = col
                        excel_earliest_year_row = row
                        print(f"Found Excel earliest year '{excel_earliest_year_str}' in header '{cell_str}' at row {row}, column {col}")
                        break
            if excel_earliest_year_col is not None:
                break
        
        if excel_earliest_year_col is None:
            print(f"Error: Could not find column with earliest year '{excel_earliest_year_str}' in Excel.")
            return False
        
        # Calculate the column shift: difference between Yahoo earliest year and Excel earliest year
        column_shift = yahoo_earliest_year - excel_earliest_year
        print(f"Column shift: {yahoo_earliest_year} - {excel_earliest_year} = {column_shift}")
        
        # Starting column for writing data = Excel earliest year column + shift
        start_col = excel_earliest_year_col + column_shift
        print(f"Starting to write data at column {start_col} (Excel earliest year column {excel_earliest_year_col} + shift {column_shift})")
        
        # Find the LTM/TTM column header
        ltm_col = None
        ltm_row = None
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value is not None:
                    cell_str = str(cell_value).strip().upper()
                    # Check for LTM or TTM
                    if cell_str in ['LTM', 'TTM', 'LTM/TTM', 'TTM/LTM']:
                        ltm_col = col
                        ltm_row = row
                        break
            if ltm_col is not None:
                break
        
        if ltm_col is None:
            print("Warning: Could not find 'LTM' or 'TTM' column header. TTM data will not be written.")
        else:
            print(f"Found 'LTM/TTM' header at row {ltm_row}, column {ltm_col}")
        
        # Find each metric and write data
        metrics_found = 0
        
        for metric in metric_order:
            # Search for metric across all columns (not just column A)
            metric_row = None
            metric_col = None
            
            # Search through all rows and columns to find the metric
            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value is not None:
                        # Check if it matches the metric (case-insensitive, strip whitespace, handle variations)
                        cell_str = str(cell_value).strip().lower()
                        metric_str = metric.lower()
                        # Exact match or match after removing spaces/underscores
                        if cell_str == metric_str or cell_str.replace(' ', '').replace('_', '') == metric_str.replace(' ', '').replace('_', ''):
                            metric_row = row
                            metric_col = col
                            break
                if metric_row is not None:
                    break
            
            if metric_row is None:
                print(f"Warning: Could not find metric '{metric}' in any column")
                continue
            
            metrics_found += 1
            print(f"Found '{metric}' at row {metric_row}, column {metric_col}")
            
            # Write data for each year sequentially starting from the calculated start column
            for year_idx, year in enumerate(years):
                # Find the corresponding year data
                value = None
                for year_data in years_data:
                    data_year = extract_year_from_date(year_data['Year'])
                    if data_year == year:
                        # Map Excel metric names to internal dictionary keys
                        dict_key = metric
                        if metric == 'LCapex':
                            dict_key = 'Capex'
                        elif metric == 'share price':
                            dict_key = 'Stock Price'
                        elif metric == 'TOCA Other Short-Term Liabilities':
                            dict_key = 'Other Short-Term Liabilities'
                        value = year_data.get(dict_key)
                        break
                
                # Calculate the column (start_col + year_idx)
                target_col = start_col + year_idx
                cell = ws.cell(row=metric_row, column=target_col)
                
                if value is not None and not pd.isna(value):
                    if isinstance(value, (int, float)):
                        # Special handling for "multiple" - it's a ratio, don't divide
                        if metric == 'multiple':
                            cell.value = value
                            cell.number_format = '#,##0.00'
                        # Special handling for "share price" - it's already in dollars, don't divide
                        elif metric == 'share price':
                            cell.value = value
                            cell.number_format = '#,##0.00'
                        else:
                            # Divide by 1000000 to match format_currency and format_number functions
                            cell.value = value / 1000000
                            cell.number_format = '#,##0.00'
                    else:
                        cell.value = value
                else:
                    cell.value = 'N/A'
            
            # Write TTM data if LTM column was found
            if ltm_col is not None:
                # Find TTM data (Year == 'LTM')
                ttm_value = None
                for year_data in years_data:
                    if year_data.get('Year') == 'LTM':
                        # Map Excel metric names to internal dictionary keys
                        dict_key = metric
                        if metric == 'LCapex':
                            dict_key = 'Capex'
                        elif metric == 'share price':
                            dict_key = 'Stock Price'
                        elif metric == 'TOCA Other Short-Term Liabilities':
                            dict_key = 'Other Short-Term Liabilities'
                        ttm_value = year_data.get(dict_key)
                        break
                
                # Write TTM value to LTM column
                ttm_cell = ws.cell(row=metric_row, column=ltm_col)
                if ttm_value is not None and not pd.isna(ttm_value):
                    if isinstance(ttm_value, (int, float)):
                        # Special handling for "multiple" - it's a ratio, don't divide
                        if metric == 'multiple':
                            ttm_cell.value = ttm_value
                            ttm_cell.number_format = '#,##0.00'
                        # Special handling for "share price" - it's already in dollars, don't divide
                        elif metric == 'share price':
                            ttm_cell.value = ttm_value
                            ttm_cell.number_format = '#,##0.00'
                        else:
                            # Divide by 1000000 to match format_currency and format_number functions
                            ttm_cell.value = ttm_value / 1000000
                            ttm_cell.number_format = '#,##0.00'
                    else:
                        ttm_cell.value = ttm_value
                else:
                    ttm_cell.value = 'N/A'
        
        # Save the file
        wb.save(file_path)
        print(f"\nSuccessfully updated {metrics_found} metrics in sheet '{sheet_name}'")
        print(f"Data saved to: {file_path}")
        return True
        
    except FileNotFoundError:
        print(f"Error: File not found: {file_path}")
        return False
    except Exception as e:
        print(f"Error writing to Excel file: {e}")
        import traceback
        traceback.print_exc()
        return False

def get_yahoo_financials(ticker, excel_path=None, sheet_name=None):
    try:
        # Create ticker object
        stock = yf.Ticker(ticker)
        
        print(f"Fetching 5 years of financial data for {ticker}...")
        
        # Get financial statements (annual data)
        financials = stock.financials  # Income statement
        cashflow = stock.cashflow      # Cash flow statement
        balance_sheet = stock.balance_sheet  # Balance sheet
        
        if financials.empty:
            print(f"Error: No financial data available for {ticker}")
            return None
        
        # Get the number of years available (up to 5)
        num_years = min(5, financials.shape[1])
        
        # Get Yahoo Finance info for EV/EBITDA multiple
        print("Fetching EV/EBITDA multiple from Yahoo Finance...")
        yahoo_multiple = None
        try:
            info = stock.info
            if 'enterpriseToEbitda' in info and info['enterpriseToEbitda'] is not None:
                yahoo_multiple = info['enterpriseToEbitda']
                print(f"Found EV/EBITDA multiple from Yahoo Finance: {yahoo_multiple}")
            elif 'enterpriseValue' in info and 'ebitda' in info:
                if info['ebitda'] is not None and info['ebitda'] != 0:
                    yahoo_multiple = info['enterpriseValue'] / info['ebitda']
                    print(f"Calculated EV/EBITDA multiple from Yahoo Finance: {yahoo_multiple}")
        except Exception as e:
            print(f"Warning: Could not get EV/EBITDA multiple from Yahoo Finance: {e}")
        
        # Get historical stock prices (starting from 2020 or the earliest year available)
        print("Fetching historical stock prices...")
        earliest_year = 2020  # Default to 2020
        if num_years > 0 and financials.columns[0] is not None:
            extracted_year = extract_year_from_date(financials.columns[0])
            if extracted_year is not None:
                # Use the earliest year from financials, or 2020, whichever is later
                earliest_year = max(extracted_year, 2020)
        
        # Try to get historical prices from the beginning if we need years before 2020
        historical_prices = {}
        try:
            # Get max history to cover all years
            hist = stock.history(period="max")
            if not hist.empty:
                hist.index = pd.to_datetime(hist.index)
                # Get all years from earliest_year onwards
                years = [y for y in hist.index.year.unique() if y >= earliest_year]
                for year in years:
                    year_data = hist[hist.index.year == year]
                    if not year_data.empty:
                        last_day = year_data.index.max()
                        price = year_data.loc[last_day, 'Close']
                        historical_prices[year] = float(price)
        except Exception as e:
            print(f"Warning: Could not fetch all historical prices: {e}")
            # Fallback to the simpler method
            historical_prices = get_historical_stock_prices(ticker, start_year=earliest_year)
        
        # Get column names (dates) for the years
        years_data = []
        
        for i in range(num_years):
            year_data = {}
            
            # Get the date for this year
            if financials.columns[i] is not None:
                year_date = financials.columns[i]
                if isinstance(year_date, pd.Timestamp):
                    year_str = year_date.strftime('%Y-%m-%d')
                else:
                    year_str = str(year_date)
            else:
                year_str = f"Year {i+1}"
            
            year_data['Year'] = year_str
            
            # Extract metrics from income statement (financials)
            year_data['Revenue'] = extract_value(financials, ['Total Revenue', 'Revenue', 'Operating Revenue'], i)
            year_data['COGS'] = extract_value(financials, ['Cost Of Revenue', 'Cost of Revenue', 'Cost Of Goods Sold', 'Cost of Goods Sold'], i)
            year_data['Gross Profit'] = extract_value(financials, ['Gross Profit'], i)
            
            # SG&A (Selling, General & Administrative)
            year_data['SG&A'] = extract_value(financials, ['Selling General And Administration', 'Selling General And Administrative', 'Selling And Marketing Expense', 'General And Administrative Expense'], i)
            
            # D&A (Depreciation & Amortization)
            # Try income statement first, then cash flow
            year_data['D&A'] = extract_value(financials, ['Depreciation And Amortization', 'Depreciation Amortization Depletion', 'Depreciation'], i)
            if year_data['D&A'] is None:
                # Try from cash flow statement
                year_data['D&A'] = extract_value(cashflow, ['Depreciation And Amortization', 'Depreciation Amortization Depletion', 'Depreciation'], i)
            
            # R&D (Research and Development)
            year_data['R&D'] = extract_value(financials, ['Research And Development', 'Research Development', 'Research Development And Engineering'], i)
            
            ebit = extract_value(financials, ['EBIT', 'Operating Income', 'Operating Profit'], i)
            year_data['EBIT'] = ebit
            year_data['EBITDA'] = extract_value(financials, ['EBITDA'], i)
            interest_expense = extract_value(financials, ['Interest Expense', 'Interest And Debt Expense'], i)
            year_data['Interest Expense'] = interest_expense
            
            # Calculate EBT = EBIT - Interest Expense
            if ebit is not None and interest_expense is not None and not pd.isna(ebit) and not pd.isna(interest_expense):
                year_data['EBT'] = ebit - interest_expense
            else:
                year_data['EBT'] = None
            
            year_data['Taxes'] = extract_value(financials, ['Tax Provision', 'Income Tax Expense', 'Taxes'], i)
            year_data['Net Income'] = extract_value(financials, ['Net Income', 'Net Income Common Stockholders'], i)
            year_data['# FDSO (millions)'] = extract_value(financials, ['Diluted Average Shares', 'Diluted Shares Outstanding'], i)
            
            # Extract metrics from cash flow statement
            year_data['Free Cash Flow'] = extract_value(cashflow, ['Free Cash Flow'], i)
            year_data['Capex'] = extract_value(cashflow, ['Capital Expenditure', 'Capital Expenditure Reported'], i)
            year_data['Purchase of Intangibles'] = extract_value(cashflow, ['Purchase Of Intangible Assets', 'Purchase Of Intangibles', 'Purchase of Intangible'], i)

            # Change in Working Capital from cash flow statement
            working_capital_change = extract_value(cashflow, ['Change In Working Capital', 'Change in Working Capital', 'Changes In Working Capital'], i)
            year_data['Working Capital Change'] = working_capital_change

            # Debt issuance and repayment for FCFE
            debt_issuance = extract_value(cashflow, ['Issuance Of Debt', 'Issuance of Debt', 'Issuance Long Term Debt', 'Issuance Of Long Term Debt', 'Issuance (Retirement) of Debt, Net'], i)
            debt_repayment = extract_value(cashflow, ['Repayment Of Debt', 'Repayments Of Debt', 'Repayment of Long Term Debt', 'Repayment Of Long Term Debt'], i)
            year_data['Debt Issuance'] = debt_issuance
            year_data['Debt Repayment'] = debt_repayment

            amortization_intangibles = extract_value(cashflow, ['Amortization Of Intangible Assets', 'Amortization Of Intangibles'], i)
            if amortization_intangibles is None:
                amortization_intangibles = extract_value(financials, ['Amortization Of Intangibles', 'Amortization of Intangibles'], i)
            year_data['Amortization Intangibles'] = amortization_intangibles

            non_operating_interest = extract_value(financials, ['Interest Expense Non Operating', 'Non Operating Interest Expense'], i)
            year_data['Non Operating Interest Expense'] = non_operating_interest

            # Extract metrics from balance sheet
            # Accounts Receivable
            year_data['accounts receivables'] = extract_value(balance_sheet, ['Accounts Receivable', 'Net Receivables', 'Trade And Other Receivables'], i)
            
            # Inventory
            year_data['inventories'] = extract_value(balance_sheet, ['Inventory', 'Inventories', 'Total Inventory'], i)
            
            # Accounts Payable
            year_data['accounts payable'] = extract_value(balance_sheet, ['Accounts Payable', 'Trade And Other Payables', 'Payables'], i)
            
            # Capture totals for working capital calculations
            total_current_assets = extract_value(balance_sheet, ['Current Assets', 'Total Current Assets'], i)
            total_current_liabilities = extract_value(balance_sheet, ['Current Liabilities', 'Total Current Liabilities'], i)
            
            # Other Current Assets (for prepaid expenses row)
            other_current_assets = extract_value(balance_sheet, ['Other Current Assets', 'Other Short Term Assets'], i)
            if other_current_assets is None:
                cash = extract_value(balance_sheet, ['Cash And Cash Equivalents', 'Cash Cash Equivalents And Short Term Investments'], i)
                inventory = year_data['inventories']
                receivables = year_data['accounts receivables']
                if total_current_assets is not None:
                    other_current_assets = total_current_assets
                    if cash is not None:
                        other_current_assets -= cash
                    if inventory is not None:
                        other_current_assets -= inventory
                    if receivables is not None:
                        other_current_assets -= receivables
            year_data['prepaid expenses'] = other_current_assets
            
            # Other Current Liabilities (Total Current Liabilities - Accounts Payable)
            accounts_payable = year_data['accounts payable']
            if total_current_liabilities is not None and accounts_payable is not None:
                year_data['Other Short-Term Liabilities'] = total_current_liabilities - accounts_payable
            elif total_current_liabilities is not None:
                year_data['Other Short-Term Liabilities'] = total_current_liabilities
            else:
                year_data['Other Short-Term Liabilities'] = extract_value(balance_sheet, ['Other Current Liabilities', 'Other Short Term Liabilities'], i)
            
            # Calculate Net Debt
            net_debt = calculate_net_debt(balance_sheet, i)
            year_data['Net Debt'] = net_debt
            
            # Working Capital level = Total Current Assets - Total Current Liabilities
            working_capital_level = None
            if total_current_assets is not None and total_current_liabilities is not None:
                working_capital_level = total_current_assets - total_current_liabilities
            year_data['Working Capital'] = working_capital_level

            # Calculate Free Cash Flows using provided formulas
            ebit_value = year_data.get('EBIT')
            d_and_a_value = year_data.get('D&A') or 0
            amort_intangibles_value = year_data.get('Amortization Intangibles') or 0
            capex_value = year_data.get('Capex') or 0
            purchase_intangibles_value = year_data.get('Purchase of Intangibles') or 0
            working_capital_change_value = year_data.get('Working Capital Change')
            non_operating_interest_value = year_data.get('Non Operating Interest Expense') or 0
            interest_expense_value = year_data.get('Interest Expense') or 0

            fcff = None
            fcfe = None
            if (ebit_value is not None and not pd.isna(ebit_value) and
                working_capital_change_value is not None and not pd.isna(working_capital_change_value)):
                tax_adjusted_ebit = ebit_value * (1 - 0.375)
                fcff = (tax_adjusted_ebit + d_and_a_value + amort_intangibles_value +
                        capex_value + purchase_intangibles_value - non_operating_interest_value -
                        working_capital_change_value)

                fcfe = (tax_adjusted_ebit + interest_expense_value + d_and_a_value +
                        amort_intangibles_value + capex_value + purchase_intangibles_value -
                        working_capital_change_value)

            year_data['FCFF'] = fcff
            year_data['FCFE'] = fcfe

            # Get year for stock price lookup
            year = extract_year_from_date(year_str)
            
            # Get stock price for this year (last trading day of the year)
            stock_price = None
            if year and year in historical_prices:
                stock_price = historical_prices[year]
            year_data['Stock Price'] = stock_price
            
            # Use EV/EBITDA multiple from Yahoo Finance
            year_data['multiple'] = yahoo_multiple
            
            years_data.append(year_data)
        
        # Calculate TTM (Trailing Twelve Months) using quarterly data
        print("Calculating TTM (Trailing Twelve Months) values...")
        ttm_data = {}
        
        try:
            # Get Yahoo Finance info for current share price
            info = stock.info
            current_share_price = None
            
            # Get current share price from Yahoo Finance
            if 'currentPrice' in info and info['currentPrice'] is not None:
                current_share_price = info['currentPrice']
            elif 'regularMarketPrice' in info and info['regularMarketPrice'] is not None:
                current_share_price = info['regularMarketPrice']
            elif 'previousClose' in info and info['previousClose'] is not None:
                current_share_price = info['previousClose']
            else:
                # Try to get from recent history
                try:
                    recent_hist = stock.history(period="1d")
                    if not recent_hist.empty:
                        current_share_price = float(recent_hist['Close'].iloc[-1])
                except:
                    pass
            
            # Use EV/EBITDA multiple from Yahoo Finance (already fetched earlier)
            ttm_data['multiple'] = yahoo_multiple
            ttm_data['Stock Price'] = current_share_price
            
            # Get quarterly financial statements for TTM calculations
            quarterly_financials = stock.quarterly_financials
            quarterly_cashflow = stock.quarterly_cashflow
            quarterly_balance_sheet = stock.quarterly_balance_sheet
            
            if not quarterly_financials.empty:
                # Get the last 4 quarters (most recent TTM)
                num_quarters = min(4, quarterly_financials.shape[1])
                
                # Helper function to sum TTM values
                def sum_ttm_quarters(df, row_names):
                    total = 0
                    has_data = False
                    for q in range(num_quarters):
                        val = extract_value(df, row_names, q)
                        if val is not None and not pd.isna(val):
                            total += val
                            has_data = True
                    return total if has_data else None
                
                # Calculate TTM for each metric (sum of last 4 quarters)
                ttm_data['Revenue'] = sum_ttm_quarters(quarterly_financials, ['Total Revenue', 'Revenue', 'Operating Revenue'])
                ttm_data['COGS'] = sum_ttm_quarters(quarterly_financials, ['Cost Of Revenue', 'Cost of Revenue', 'Cost Of Goods Sold', 'Cost of Goods Sold'])
                ttm_data['Gross Profit'] = sum_ttm_quarters(quarterly_financials, ['Gross Profit'])
                
                # SG&A TTM
                ttm_data['SG&A'] = sum_ttm_quarters(quarterly_financials, ['Selling General And Administration', 'Selling General And Administrative', 'Selling And Marketing Expense', 'General And Administrative Expense'])
                
                # D&A TTM (try income statement first, then cash flow)
                ttm_data['D&A'] = sum_ttm_quarters(quarterly_financials, ['Depreciation And Amortization', 'Depreciation Amortization Depletion', 'Depreciation'])
                if ttm_data['D&A'] is None and not quarterly_cashflow.empty:
                    ttm_data['D&A'] = sum_ttm_quarters(quarterly_cashflow, ['Depreciation And Amortization', 'Depreciation Amortization Depletion', 'Depreciation'])
                
                # R&D TTM
                ttm_data['R&D'] = sum_ttm_quarters(quarterly_financials, ['Research And Development', 'Research Development', 'Research Development And Engineering'])
                
                ebit_ttm = sum_ttm_quarters(quarterly_financials, ['EBIT', 'Operating Income', 'Operating Profit'])
                ttm_data['EBITDA'] = sum_ttm_quarters(quarterly_financials, ['EBITDA'])
                interest_expense_ttm = sum_ttm_quarters(quarterly_financials, ['Interest Expense', 'Interest And Debt Expense'])
                
                # Calculate EBT TTM
                if ebit_ttm is not None and interest_expense_ttm is not None:
                    ttm_data['EBT'] = ebit_ttm - interest_expense_ttm
                else:
                    ttm_data['EBT'] = None
                
                ttm_data['Taxes'] = sum_ttm_quarters(quarterly_financials, ['Tax Provision', 'Income Tax Expense', 'Taxes'])
                ttm_data['Net Income'] = sum_ttm_quarters(quarterly_financials, ['Net Income', 'Net Income Common Stockholders'])
                ttm_data['# FDSO (millions)'] = extract_value(quarterly_financials, ['Diluted Average Shares', 'Diluted Shares Outstanding'], 0)
                ttm_data['Interest Expense'] = interest_expense_ttm
                
                # Get TTM from cash flow
                if not quarterly_cashflow.empty:
                    ttm_data['Free Cash Flow'] = sum_ttm_quarters(quarterly_cashflow, ['Free Cash Flow'])
                    ttm_data['Capex'] = sum_ttm_quarters(quarterly_cashflow, ['Capital Expenditure', 'Capital Expenditure Reported'])
                    working_capital_change_ttm = sum_ttm_quarters(quarterly_cashflow, ['Change In Working Capital', 'Change in Working Capital', 'Changes In Working Capital'])
                    ttm_data['Working Capital Change'] = working_capital_change_ttm
                    debt_issuance_ttm = sum_ttm_quarters(quarterly_cashflow, ['Issuance Of Debt', 'Issuance of Debt', 'Issuance Long Term Debt', 'Issuance Of Long Term Debt', 'Issuance (Retirement) of Debt, Net'])
                    debt_repayment_ttm = sum_ttm_quarters(quarterly_cashflow, ['Repayment Of Debt', 'Repayments Of Debt', 'Repayment of Long Term Debt', 'Repayment Of Long Term Debt'])
                    ttm_data['Debt Issuance'] = debt_issuance_ttm
                    ttm_data['Debt Repayment'] = debt_repayment_ttm
                    amort_intangibles_ttm = sum_ttm_quarters(quarterly_cashflow, ['Amortization Of Intangible Assets', 'Amortization Of Intangibles'])
                    if amort_intangibles_ttm is None:
                        amort_intangibles_ttm = sum_ttm_quarters(quarterly_financials, ['Amortization Of Intangibles', 'Amortization of Intangibles'])
                    ttm_data['Amortization Intangibles'] = amort_intangibles_ttm
                    purchase_intangibles_ttm = sum_ttm_quarters(quarterly_cashflow, ['Purchase Of Intangible Assets', 'Purchase Of Intangibles', 'Purchase of Intangible'])
                    ttm_data['Purchase of Intangibles'] = purchase_intangibles_ttm
                else:
                    ttm_data['Working Capital Change'] = None
                    ttm_data['Debt Issuance'] = None
                    ttm_data['Debt Repayment'] = None
                    ttm_data['Amortization Intangibles'] = None
                    ttm_data['Purchase of Intangibles'] = None

                # Get TTM from balance sheet (use most recent)
                if not quarterly_balance_sheet.empty:
                    ttm_other_st_liab = extract_value(quarterly_balance_sheet, ['Current Liabilities', 'Total Current Liabilities'], 0)
                    ttm_data['Other Short-Term Liabilities'] = ttm_other_st_liab
                    ttm_net_debt = calculate_net_debt(quarterly_balance_sheet, 0)
                    ttm_data['Net Debt'] = ttm_net_debt
                    working_capital_ttm_level = extract_value(quarterly_balance_sheet, ['Working Capital'], 0)
                    if working_capital_ttm_level is None:
                        ttm_current_assets = extract_value(quarterly_balance_sheet, ['Current Assets', 'Total Current Assets'], 0)
                        ttm_current_liabilities = extract_value(quarterly_balance_sheet, ['Current Liabilities', 'Total Current Liabilities'], 0)
                        if ttm_current_assets is not None and ttm_current_liabilities is not None:
                            working_capital_ttm_level = ttm_current_assets - ttm_current_liabilities
                    ttm_data['Working Capital'] = working_capital_ttm_level
                    non_operating_interest_ttm = extract_value(quarterly_financials, ['Interest Expense Non Operating', 'Non Operating Interest Expense'], 0)
                    ttm_data['Non Operating Interest Expense'] = non_operating_interest_ttm
                else:
                    ttm_data['Other Short-Term Liabilities'] = None
                    ttm_data['Net Debt'] = None
                    ttm_data['Working Capital'] = None
                    ttm_data['Non Operating Interest Expense'] = None

                # Calculate TTM FCFF and FCFE if components are available
                ebit_ttm_value = ebit_ttm
                d_and_a_ttm = ttm_data.get('D&A') or 0
                amort_intangibles_ttm = ttm_data.get('Amortization Intangibles') or 0
                capex_ttm = ttm_data.get('Capex') or 0
                purchase_intangibles_ttm = ttm_data.get('Purchase of Intangibles') or 0
                working_capital_change_ttm = ttm_data.get('Working Capital Change')
                non_operating_interest_ttm = ttm_data.get('Non Operating Interest Expense') or 0
                interest_expense_ttm = ttm_data.get('Interest Expense') or 0

                if (ebit_ttm_value is not None and not pd.isna(ebit_ttm_value) and
                    working_capital_change_ttm is not None and not pd.isna(working_capital_change_ttm)):
                    tax_adjusted_ebit_ttm = ebit_ttm_value * (1 - 0.375)
                    ttm_data['FCFF'] = (tax_adjusted_ebit_ttm + d_and_a_ttm + amort_intangibles_ttm +
                                        capex_ttm + purchase_intangibles_ttm - non_operating_interest_ttm -
                                        working_capital_change_ttm)
                else:
                    ttm_data['FCFF'] = None

                issuance_ttm = ttm_data.get('Debt Issuance')
                repayment_ttm = ttm_data.get('Debt Repayment')
                issuance_ttm_available = issuance_ttm is not None and not pd.isna(issuance_ttm)
                repayment_ttm_available = repayment_ttm is not None and not pd.isna(repayment_ttm)
                if (ttm_data['FCFF'] is not None and (issuance_ttm_available or repayment_ttm_available or interest_expense_ttm != 0)):
                    issuance_component_ttm = issuance_ttm if issuance_ttm_available else 0
                    repayment_component_ttm = repayment_ttm if repayment_ttm_available else 0
                    ttm_data['FCFE'] = (tax_adjusted_ebit_ttm + interest_expense_ttm + d_and_a_ttm +
                                         amort_intangibles_ttm + capex_ttm + purchase_intangibles_ttm -
                                         working_capital_change_ttm + issuance_component_ttm + repayment_component_ttm)
                else:
                    ttm_data['FCFE'] = None
        
        except Exception as e:
            print(f"Warning: Could not calculate TTM values: {e}")
            import traceback
            traceback.print_exc()
        
        # Ensure all expected keys exist in TTM data
        for key in ['Revenue', 'COGS', 'Gross Profit', 'SG&A', 'D&A', 'R&D', 'EBT', 'EBITDA', 'Net Income',
                    'Free Cash Flow', 'FCFF', 'FCFE', 'Capex', 'Net Debt', 'Other Short-Term Liabilities',
                    'Taxes', 'Interest Expense', '# FDSO (millions)', 'Working Capital', 'Working Capital Change',
                    'Debt Issuance', 'Debt Repayment', 'Amortization Intangibles', 'Purchase of Intangibles',
                    'Non Operating Interest Expense']:
            ttm_data.setdefault(key, None)
        
        # Add TTM to years_data
        ttm_data['Year'] = 'LTM'
        years_data.append(ttm_data)
        
        # Display results in a table format
        print(f"\n{'='*150}")
        print(f"5-Year Financial Data for {ticker}")
        print(f"{'='*150}\n")
        
        # Print header with better spacing
        header_format = "{:<12} {:<18} {:<18} {:<18} {:<18} {:<18} {:<18} {:<18} {:<18} {:<18} {:<18} {:<18} {:<18} {:<18} {:<18} {:<18} {:<18} {:<18}"
        print(header_format.format(
            'Year', 'Revenue', 'COGS', 'Gross Profit', 'EBT', 'EBITDA', 'Net Income',
            'Free Cash Flow', 'FCFF', 'FCFE', 'Capex', 'Net Debt', 'Other ST Liab', 'Taxes', 
            'Interest Exp', '# FDSO (millions)', 'multiple', 'Stock Price'
        ))
        print("-" * 150)
        
        # Print data for each year
        for year_data in years_data:
            # Format multiple (ratio)
            multiple_str = f"{year_data.get('multiple', 'N/A'):.2f}" if year_data.get('multiple') is not None and not pd.isna(year_data.get('multiple')) else 'N/A'
            # Format stock price (currency)
            stock_price_str = f"${year_data.get('Stock Price', 'N/A'):.2f}" if year_data.get('Stock Price') is not None and not pd.isna(year_data.get('Stock Price')) else 'N/A'
            
            print(header_format.format(
                year_data['Year'][:12],
                format_currency(year_data['Revenue']),
                format_currency(year_data['COGS']),
                format_currency(year_data['Gross Profit']),
                format_currency(year_data['EBT']),
                format_currency(year_data['EBITDA']),
                format_currency(year_data['Net Income']),
                format_currency(year_data['Free Cash Flow']),
                format_currency(year_data['FCFF']),
                format_currency(year_data['FCFE']),
                format_currency(year_data['Capex']),
                format_currency(year_data['Net Debt']),
                format_currency(year_data['Other Short-Term Liabilities']),
                format_currency(year_data['Taxes']),
                format_currency(year_data['Interest Expense']),
                format_number(year_data['# FDSO (millions)']),
                multiple_str,
                stock_price_str
            ))
        
        print(f"\n{'='*150}\n")
        
        # Write to Excel if path and sheet name provided
        if excel_path and sheet_name:
            write_to_excel(years_data, excel_path, sheet_name)
        
        # Return the data as a list of dictionaries
        return years_data
        
    except Exception as e:
        print(f"Error fetching data: {e}")
        print("Please check that the ticker symbol is correct.")
        import traceback
        traceback.print_exc()
        return None

class DataFetchThread(QThread):
    """Thread for fetching financial data without freezing the UI"""
    finished = pyqtSignal(bool, str, str)  # success, output, error_output
    log_message = pyqtSignal(str)
    
    def __init__(self, ticker, excel_path, sheet_name):
        super().__init__()
        self.ticker = ticker
        self.excel_path = excel_path
        self.sheet_name = sheet_name
    
    def run(self):
        """Run the data fetching in background thread"""
        import sys
        from io import StringIO
        
        try:
            self.log_message.emit(f"Fetching financial data for {self.ticker}...")
            
            # Redirect print statements to capture output
            old_stdout = sys.stdout
            old_stderr = sys.stderr
            string_stdout = StringIO()
            string_stderr = StringIO()
            sys.stdout = string_stdout
            sys.stderr = string_stderr
            
            try:
                result = get_yahoo_financials(self.ticker, self.excel_path, self.sheet_name)
                
                # Get captured output
                output = string_stdout.getvalue()
                error_output = string_stderr.getvalue()
                
                # Restore stdout/stderr
                sys.stdout = old_stdout
                sys.stderr = old_stderr
                
                self.finished.emit(bool(result), output, error_output)
                
            except Exception as e:
                # Restore stdout/stderr even on error
                sys.stdout = old_stdout
                sys.stderr = old_stderr
                error_output = f"Error: {str(e)}\n"
                import traceback
                error_output += traceback.format_exc()
                self.finished.emit(False, "", error_output)
                
        except Exception as e:
            error_output = f"Error: {str(e)}\n"
            import traceback
            error_output += traceback.format_exc()
            self.finished.emit(False, "", error_output)


class FinancialDataGUI(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("GIR DCF Historical Autocomplete")
        self.setGeometry(100, 100, 1100, 750)
        self.setStyleSheet("background-color: black; color: white;")

        # Metric mapping reference (Yahoo Finance key -> Excel row label)
        self.metric_mapping = [
            ("Total Revenue", "Revenue"),
            ("Cost of Revenue", "COGS"),
            ("Gross Profit", "Gross Profit"),
            ("Selling, General & Administrative", "SG&A"),
            ("Depreciation & Amortization", "D&A"),
            ("Research & Development", "R&D"),
            ("EBIT (less Interest Expense)", "EBT"),
            ("EBITDA", "EBITDA"),
            ("Net Income", "Net Income"),
            ("Free Cash Flow", "Free Cash Flow"),
            ("Unlevered Free Cash Flow", "FCFF"),
            ("Levered Free Cash Flow", "FCFE"),
            ("Capital Expenditure", "LCapex"),
            ("Net Debt", "Net Debt"),
            ("Accounts Receivable", "accounts receivables"),
            ("Other Current Assets", "prepaid expenses"),
            ("Inventory", "inventories"),
            ("Accounts Payable", "accounts payable"),
            ("Other Short-Term Liabilities", "TOCA Other Short-Term Liabilities"),
            ("Tax Provision", "Taxes"),
            ("Interest Expense", "Interest Expense"),
            ("Diluted Average Shares", "# FDSO (millions)"),
            ("EV/EBITDA Multiple", "multiple"),
            ("Year-End Share Price", "share price")
        ]

        # Central widget
        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        # Main layout
        main_layout = QVBoxLayout()
        central_widget.setLayout(main_layout)
        main_layout.setSpacing(20)
        main_layout.setContentsMargins(20, 20, 20, 20)

        # Top section with logo and header
        top_section_layout = QHBoxLayout()
        top_section_layout.setSpacing(30)
        top_section_layout.setAlignment(Qt.AlignLeft)

        # Logo (load from provided path)
        logo_label = QLabel()
        logo_path = "/Users/timothyh/Desktop/ICDCFAUTO/IC logo.png"
        if os.path.exists(logo_path):
            pixmap = QPixmap(logo_path)
            if not pixmap.isNull():
                logo_label.setPixmap(pixmap.scaled(180, 180, Qt.KeepAspectRatio, Qt.SmoothTransformation))
                logo_label.setFixedSize(200, 200)
            else:
                logo_label.setText("IC Logo")
                logo_label.setStyleSheet("font-size: 24px; font-weight: bold; color: white;")
                logo_label.setAlignment(Qt.AlignCenter)
                logo_label.setFixedSize(200, 200)
        else:
            logo_label.setText("IC Logo")
            logo_label.setStyleSheet("font-size: 24px; font-weight: bold; color: white;")
            logo_label.setAlignment(Qt.AlignCenter)
            logo_label.setFixedSize(200, 200)
        top_section_layout.addWidget(logo_label)

        header_container = QVBoxLayout()
        header_container.setSpacing(10)
        header_container.setContentsMargins(0, 30, 0, 0)

        header_label = QLabel("GIR DCF Autocomplete")
        header_label.setStyleSheet("font-size: 28px; font-weight: bold; color: white;")
        header_container.addWidget(header_label)

        description_label = QLabel(
            "Automated tool that pulls 5Y + LTM historical metrics from Yahoo Finance and"
            " writes them directly into your DCF workbook. Provide a ticker, workbook path, and"
            " sheet, the tool handles the rest."
        )
        description_label.setWordWrap(True)
        description_label.setStyleSheet("font-size: 13px; color: #cccccc; line-height: 150%;")
        header_container.addWidget(description_label)
        header_container.addStretch()

        top_section_layout.addLayout(header_container)
        main_layout.addLayout(top_section_layout)

        # Ticker input
        ticker_layout = QHBoxLayout()
        ticker_label = QLabel("Ticker Symbol:")
        ticker_label.setStyleSheet("font-size: 12px; color: white;")
        ticker_label.setFixedWidth(120)
        ticker_layout.addWidget(ticker_label)

        self.ticker_entry = QLineEdit()
        self.ticker_entry.setStyleSheet("font-size: 12px; background-color: white; color: black; padding: 5px;")
        self.ticker_entry.setFixedWidth(200)
        ticker_layout.addWidget(self.ticker_entry)
        ticker_layout.addStretch()
        main_layout.addLayout(ticker_layout)

        # Excel file path input
        file_layout = QHBoxLayout()
        file_label = QLabel("Excel File Path:")
        file_label.setStyleSheet("font-size: 12px; color: white;")
        file_label.setFixedWidth(120)
        file_layout.addWidget(file_label)

        self.file_path_entry = QLineEdit()
        self.file_path_entry.setStyleSheet("font-size: 12px; background-color: white; color: black; padding: 5px;")
        file_layout.addWidget(self.file_path_entry)

        browse_button = QPushButton("Browse")
        browse_button.setStyleSheet("background-color: #0066cc; color: white; padding: 5px 15px; font-size: 12px;")
        browse_button.clicked.connect(self.browse_file)
        browse_button.setFixedWidth(100)
        file_layout.addWidget(browse_button)
        main_layout.addLayout(file_layout)

        # Sheet name input
        sheet_layout = QHBoxLayout()
        sheet_label = QLabel("Sheet Name:")
        sheet_label.setStyleSheet("font-size: 12px; color: white;")
        sheet_label.setFixedWidth(120)
        sheet_layout.addWidget(sheet_label)

        self.sheet_entry = QLineEdit()
        self.sheet_entry.setStyleSheet("font-size: 12px; background-color: white; color: black; padding: 5px;")
        self.sheet_entry.setFixedWidth(300)
        sheet_layout.addWidget(self.sheet_entry)
        sheet_layout.addStretch()
        main_layout.addLayout(sheet_layout)

        # Submit button
        self.submit_button = QPushButton("Submit")
        self.submit_button.setStyleSheet("""
            background-color: #0066cc; 
            color: white; 
            padding: 10px 30px; 
            font-size: 14px; 
            font-weight: bold;
            border-radius: 5px;
        """)
        self.submit_button.setFixedHeight(40)
        self.submit_button.clicked.connect(self.on_submit)
        main_layout.addWidget(self.submit_button)

        # Console + mapping layout
        console_table_layout = QHBoxLayout()
        console_table_layout.setSpacing(20)

        # Status/Output area
        status_container = QVBoxLayout()
        status_label = QLabel("Status:")
        status_label.setStyleSheet("font-size: 12px; color: white;")
        status_container.addWidget(status_label)

        self.status_text = QTextEdit()
        self.status_text.setStyleSheet("""
            background-color: #1a1a1a; 
            color: #00ff00; 
            font-family: 'Courier New', monospace; 
            font-size: 10px;
            border: 2px solid #333;
        """)
        self.status_text.setReadOnly(True)
        status_container.addWidget(self.status_text)
        console_table_layout.addLayout(status_container, stretch=2)

        # Metric mapping table on the right
        mapping_container = QVBoxLayout()
        mapping_label = QLabel("Metric Mapping Reference")
        mapping_label.setStyleSheet("font-size: 12px; color: white; font-weight: bold;")
        mapping_container.addWidget(mapping_label)

        self.metric_table = QTableWidget()
        self.metric_table.setColumnCount(2)
        self.metric_table.setHorizontalHeaderLabels(["Yahoo Finance Metric", "Excel Row Label"])
        self.metric_table.verticalHeader().setVisible(False)
        self.metric_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.metric_table.setSelectionMode(QAbstractItemView.NoSelection)
        self.metric_table.setFocusPolicy(Qt.NoFocus)
        self.metric_table.setStyleSheet("""
            QTableWidget {
                background-color: #1a1a1a;
                color: white;
                gridline-color: #333;
                font-size: 11px;
            }
            QHeaderView::section {
                background-color: #333;
                color: white;
                font-weight: bold;
                border: 1px solid #444;
            }
        """)
        self.metric_table.horizontalHeader().setStretchLastSection(True)
        self.metric_table.horizontalHeader().setDefaultSectionSize(200)
        self.populate_metric_mapping_table()
        mapping_container.addWidget(self.metric_table)

        console_table_layout.addLayout(mapping_container, stretch=1)
        main_layout.addLayout(console_table_layout)

        # Footer hyperlink
        footer_label = QLabel()
        footer_label.setText('<span style="color:#888">Created by </span><a href="https://www.linkedin.com/in/timothyhawks/" style="color:#4da6ff;">Timothy Hawks</a>')
        footer_label.setTextFormat(Qt.RichText)
        footer_label.setTextInteractionFlags(Qt.TextBrowserInteraction)
        footer_label.setOpenExternalLinks(True)
        footer_label.setAlignment(Qt.AlignCenter)
        footer_label.setStyleSheet("font-size: 11px; padding-top: 10px;")
        main_layout.addWidget(footer_label)

        # Thread for data fetching
        self.fetch_thread = None

    def log_message(self, message):
        """Add message to status area"""
        self.status_text.append(message)
        # Auto-scroll to bottom
        scrollbar = self.status_text.verticalScrollBar()
        scrollbar.setValue(scrollbar.maximum())
    
    def browse_file(self):
        """Browse for Excel file"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Select Excel File",
            "",
            "Excel files (*.xlsx *.xls);;All files (*.*)"
        )
        if file_path:
            self.file_path_entry.setText(file_path)
    
    def on_submit(self):
        """Handle submit button click"""
        # Get ticker
        ticker = self.ticker_entry.text().strip().upper()
        if not ticker:
            QMessageBox.critical(self, "Error", "Please enter a ticker symbol")
            return
        
        # Get Excel file path
        excel_path = self.file_path_entry.text().strip()
        if excel_path:
            # Validate file exists
            if not os.path.exists(excel_path):
                QMessageBox.critical(self, "Error", f"File not found: {excel_path}")
                return
            if not excel_path.lower().endswith(('.xlsx', '.xls')):
                QMessageBox.critical(self, "Error", "Please provide an Excel file (.xlsx or .xls)")
                return
        
        # Get sheet name (required only if Excel file is provided)
        sheet_name = self.sheet_entry.text().strip() if excel_path else None
        if excel_path and not sheet_name:
            QMessageBox.critical(self, "Error", "Please enter a sheet name")
            return
        
        # Clear status area
        self.status_text.clear()
        
        # Disable submit button during processing
        self.submit_button.setEnabled(False)
        self.submit_button.setText("Processing...")
        
        # Create and start the data fetching thread
        self.fetch_thread = DataFetchThread(ticker, excel_path if excel_path else None, sheet_name)
        self.fetch_thread.log_message.connect(self.log_message)
        self.fetch_thread.finished.connect(self.on_fetch_finished)
        self.fetch_thread.start()
    
    def on_fetch_finished(self, success, output, error_output):
        """Handle completion of data fetching"""
        # Re-enable submit button
        self.submit_button.setEnabled(True)
        self.submit_button.setText("Submit")
        
        # Display output
        if output:
            self.log_message(output)
        if error_output:
            self.log_message(f"Errors:\n{error_output}")
        
        if success:
            self.log_message("\n✓ Data fetched successfully!")
            if self.file_path_entry.text().strip():
                self.log_message(f"✓ Data saved to Excel file")
            QMessageBox.information(self, "Success", "Financial data fetched successfully!")
        else:
            QMessageBox.critical(self, "Error", "Failed to fetch financial data. Check the status area for details.")

    def populate_metric_mapping_table(self):
        """Populate the static metric mapping table."""
        self.metric_table.setRowCount(len(self.metric_mapping))
        for row_idx, (yahoo_metric, excel_label) in enumerate(self.metric_mapping):
            yahoo_item = QTableWidgetItem(yahoo_metric)
            excel_item = QTableWidgetItem(excel_label)
            yahoo_item.setFlags(Qt.ItemIsEnabled)
            excel_item.setFlags(Qt.ItemIsEnabled)
            self.metric_table.setItem(row_idx, 0, yahoo_item)
            self.metric_table.setItem(row_idx, 1, excel_item)
        self.metric_table.resizeRowsToContents()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = FinancialDataGUI()
    window.show()
    sys.exit(app.exec_())